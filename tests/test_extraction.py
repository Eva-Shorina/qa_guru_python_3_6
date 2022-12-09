import shutil
from zipfile import ZipFile
import csv
import os
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from zippack import creating_arch

creating_arch()
this_dir = os.path.dirname(os.path.abspath(__file__))
root2 = os.path.dirname(this_dir)
path = os.path.join(root2, 'resources')
arch_path = os.path.join(path,'archive.zip')


def test_csv():
    file = zip_.extract(zip_.namelist()[1])

    with open(file) as csvfile:
        csvfile = csv.reader(csvfile)

        for r in csvfile:
            temp = r[0]

        value = temp.split(';')[3]
        assert value == 'CCCC'


def test_pdf():
    file = zip_.extract(zip_.namelist()[3])

    reader = PdfReader(file)
    page = reader.pages[0]
    text = page.extract_text()
    print(text)
    assert 'ура' in text


def test_xlsx():
    file = zip_.extract(zip_.namelist()[4])
    workbook = load_workbook(file)
    sheet = workbook.active
    assert sheet.cell(row=3, column=2).value == 'Eagle'

def test_jpg():
    file = zip_.extract(zip_.namelist()[2])
    size = os.path.getsize(file)
    assert size == 2435

